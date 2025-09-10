# app_lookup_open_gui.py
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Iterable, List, Optional

import polars as pl

DEFAULT_OPEN_PATH = Path("data/ddbb_scd/open/open_intervals.parquet")
DEFAULT_HISTORY_GLOB = "data/ddbb_scd/history/closed_until_*.parquet"


# =========== Core helpers ===========

def _normalize_cnpj(text: str) -> Optional[str]:
    digits = re.sub(r"\D+", "", text or "")
    if len(digits) >= 14:
        digits = digits[-14:]
    digits = digits.zfill(14)
    return digits if len(digits) == 14 else None

def _unique_cnpjs(items: Iterable[str]) -> List[str]:
    out, seen = [], set()
    for t in items:
        c = _normalize_cnpj(t)
        if c and c not in seen:
            seen.add(c)
            out.append(c)
    return out

def _validate_as_of(as_of: Optional[str]) -> Optional[str]:
    if not as_of:
        return None
    as_of = as_of.strip()
    if not re.fullmatch(r"20\d{2}-\d{2}", as_of):
        raise ValueError("as-of must be 'YYYY-MM' starting with '20'. Example: 2024-11")
    return as_of


# =========== Lookups ===========

def lookup_cnae_open_many(cnpjs: List[str], open_path: Path) -> pl.DataFrame:
    if not cnpjs:
        return pl.DataFrame({"cnpj": [], "cnaes": []})
    if not open_path.exists() or open_path.stat().st_size == 0:
        raise FileNotFoundError(f"Open intervals Parquet not found or empty: {open_path}")

    lf = (
        pl.scan_parquet(open_path)
        .filter(pl.col("cnpj").is_in(cnpjs))
        .select("cnpj", "cnae")
        .group_by("cnpj")
        .agg(pl.col("cnae").unique().sort().alias("cnaes"))
    )
    present = lf.collect(streaming=True)
    missing = [c for c in cnpjs if c not in set(present.get_column("cnpj").to_list())]
    if missing:
        df_missing = pl.DataFrame({"cnpj": missing, "cnaes": [[] for _ in missing]})
        return pl.concat([present, df_missing], how="diagonal_relaxed").sort("cnpj")
    return present.sort("cnpj")


def lookup_cnae_as_of_many(
    cnpjs: List[str],
    as_of: str,
    *,
    open_path: Path = DEFAULT_OPEN_PATH,
    history_glob: str = DEFAULT_HISTORY_GLOB,
) -> pl.DataFrame:
    """
    As-of lookup using validity windows:
      valid_from_month <= as_of <= COALESCE(valid_to_month, '∞')
    Implements this by unioning:
      - open rows with valid_from_month <= as_of
      - closed rows with valid_from_month <= as_of <= valid_to_month
    Returns DF: [cnpj, cnaes(List[str])]
    """
    if not cnpjs:
        return pl.DataFrame({"cnpj": [], "cnaes": []})
    if not open_path.exists() or open_path.stat().st_size == 0:
        raise FileNotFoundError(f"Open intervals Parquet not found or empty: {open_path}")

    # Open (current) intervals that already existed by 'as_of'
    open_lf = (
        pl.scan_parquet(open_path)
        .filter(
            pl.col("cnpj").is_in(cnpjs)
            & (pl.col("valid_from_month") <= as_of)
        )
        .select("cnpj", "cnae")
    )

    # Closed intervals valid at 'as_of'
    hist_lf = (
        pl.scan_parquet(history_glob)
        .filter(
            pl.col("cnpj").is_in(cnpjs)
            & (pl.col("valid_from_month") <= as_of)
            & (as_of <= pl.col("valid_to_month"))
        )
        .select("cnpj", "cnae")
    )

    # Union and aggregate
    agg = (
        pl.concat([open_lf, hist_lf], how="vertical_relaxed")
        .group_by("cnpj")
        .agg(pl.col("cnae").unique().sort().alias("cnaes"))
        .collect(streaming=True)
    )

    # Ensure all requested CNPJs appear
    have = set(agg.get_column("cnpj").to_list())
    missing = [c for c in cnpjs if c not in have]
    if missing:
        df_missing = pl.DataFrame({"cnpj": missing, "cnaes": [[] for _ in missing]})
        agg = pl.concat([agg, df_missing], how="diagonal_relaxed")
    return agg.sort("cnpj")


# =========== File readers / writers ===========

def _sniff_delimiter(sample: str) -> str:
    return ";" if sample.count(";") > sample.count(",") else ","

def read_cnpjs_from_csv(path: Path, *, sep: Optional[str] = None, cnpj_col: Optional[str] = None) -> List[str]:
    if sep is None:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            head = f.read(4096)
        sep = _sniff_delimiter(head)
    df = pl.read_csv(path, separator=sep, infer_schema_length=50, ignore_errors=True, try_parse_dates=False)
    if cnpj_col:
        if cnpj_col not in df.columns:
            raise ValueError(f"Column '{cnpj_col}' not in {path.name}. Found: {df.columns}")
        return _unique_cnpjs(df[cnpj_col].cast(pl.Utf8, strict=False).fill_null("").to_list())
    best_name, best_hits = None, -1
    for col in df.columns:
        s = df[col].cast(pl.Utf8, strict=False).fill_null("")
        normalized = s.str.replace_all(r"\D+", "").str.slice(-14).str.zfill(14)
        hits = int((normalized.str.len_chars() == 14).sum())
        if hits > best_hits:
            best_hits, best_name = hits, col
    if not best_name:
        return []
    return _unique_cnpjs(df[best_name].cast(pl.Utf8, strict=False).fill_null("").to_list())

def read_cnpjs_from_xlsx(path: Path, *, sheet: Optional[str] = None, cnpj_col: Optional[str] = None) -> List[str]:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("Reading .xlsx requires 'openpyxl'. Install with: pip install openpyxl")
    wb = openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    columns = {h: [] for h in headers}
    for row in rows:
        for i, h in enumerate(headers):
            val = "" if row is None or i >= len(row) or row[i] is None else str(row[i])
            columns[h].append(val)
    if cnpj_col:
        if cnpj_col not in columns:
            raise ValueError(f"Column '{cnpj_col}' not in {path.name}. Found: {list(columns.keys())}")
        return _unique_cnpjs(columns[cnpj_col])
    best_name, best_hits = None, -1
    for name, vals in columns.items():
        hits = sum(1 for v in vals if _normalize_cnpj(v))
        if hits > best_hits:
            best_hits, best_name = hits, name
    return _unique_cnpjs(columns.get(best_name, []))

def export_results_csv(df: pl.DataFrame, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    rows = [{"cnpj": r["cnpj"], "cnaes": ",".join(r["cnaes"])} for r in df.iter_rows(named=True)]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["cnpj", "cnaes"])
        w.writeheader()
        w.writerows(rows)


# =========== Tkinter GUI ===========

def run_gui(default_open_path: Path = DEFAULT_OPEN_PATH, history_glob: str = DEFAULT_HISTORY_GLOB) -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    root.title("CNAE Lookup (open_intervals + as-of history)")

    # Vars
    open_path_var = tk.StringVar(value=str(default_open_path))
    as_of_var = tk.StringVar(value="")  # YYYY-MM or empty for "current"
    file_path_var = tk.StringVar(value="")
    csv_sep_var = tk.StringVar(value="auto")
    cnpj_col_var = tk.StringVar(value="")
    xlsx_sheet_var = tk.StringVar(value="")

    # Top controls
    frm_top = ttk.Frame(root); frm_top.grid(row=0, column=0, columnspan=2, sticky="we"); frm_top.columnconfigure(1, weight=1)
    ttk.Label(frm_top, text="open_intervals.parquet:").grid(row=0, column=0, padx=6, pady=4, sticky="w")
    ent_open = ttk.Entry(frm_top, textvariable=open_path_var); ent_open.grid(row=0, column=1, padx=6, pady=4, sticky="we")
    ttk.Button(frm_top, text="Browse...", command=lambda: _browse_file(open_path_var, ("Parquet", "*.parquet"))).grid(row=0, column=2, padx=6, pady=4)
    ttk.Label(frm_top, text="As of (YYYY-MM):").grid(row=1, column=0, padx=6, pady=2, sticky="w")
    ent_asof = ttk.Entry(frm_top, textvariable=as_of_var, width=12); ent_asof.grid(row=1, column=1, padx=6, pady=2, sticky="w")

    # Notebook
    nb = ttk.Notebook(root); nb.grid(row=1, column=0, columnspan=2, sticky="nsew"); root.rowconfigure(1, weight=1)

    # Tab: Text
    tab_text = ttk.Frame(nb); nb.add(tab_text, text="Paste CNPJs / free text")
    tab_text.columnconfigure(0, weight=1); tab_text.rowconfigure(0, weight=1)
    txt = tk.Text(tab_text, height=8, wrap="word"); txt.grid(row=0, column=0, sticky="nsew", padx=6, pady=6)
    ttk.Label(tab_text, text="Paste many CNPJs; separators don't matter. Example: 12.345.678/0001-95, 11.027.469/0001-30").grid(row=1, column=0, sticky="w", padx=6, pady=(0,6))

    # Tab: File
    tab_file = ttk.Frame(nb); nb.add(tab_file, text="CSV / XLSX file")
    for c in range(3): tab_file.columnconfigure(c, weight=1)
    file_path_entry = ttk.Entry(tab_file, textvariable=file_path_var); file_path_entry.grid(row=0, column=1, padx=6, pady=4, sticky="we")
    ttk.Label(tab_file, text="Input file:").grid(row=0, column=0, padx=6, pady=4, sticky="w")
    ttk.Button(tab_file, text="Browse...", command=lambda: _browse_file(file_path_var, (("CSV/XLSX", "*.csv *.xlsx"), ("CSV", "*.csv"), ("XLSX", "*.xlsx")))).grid(row=0, column=2, padx=6, pady=4)
    ttk.Label(tab_file, text="CNPJ column (optional):").grid(row=1, column=0, padx=6, pady=2, sticky="w")
    ttk.Entry(tab_file, textvariable=cnpj_col_var).grid(row=1, column=1, padx=6, pady=2, sticky="we")
    ttk.Label(tab_file, text="XLSX sheet (optional):").grid(row=2, column=0, padx=6, pady=2, sticky="w")
    ttk.Entry(tab_file, textvariable=xlsx_sheet_var).grid(row=2, column=1, padx=6, pady=2, sticky="we")
    ttk.Label(tab_file, text="CSV delimiter:").grid(row=3, column=0, padx=6, pady=2, sticky="w")
    cmb_sep = ttk.Combobox(tab_file, textvariable=csv_sep_var, values=["auto", ",", ";"], state="readonly", width=8); cmb_sep.current(0)
    cmb_sep.grid(row=3, column=1, padx=6, pady=2, sticky="w")

    # Actions
    frm_actions = ttk.Frame(root); frm_actions.grid(row=2, column=0, columnspan=2, sticky="we")
    ttk.Button(frm_actions, text="Lookup", command=lambda: _do_lookup()).grid(row=0, column=0, padx=6, pady=6, sticky="w")
    ttk.Button(frm_actions, text="Export CSV...", command=lambda: _export_csv()).grid(row=0, column=1, padx=6, pady=6, sticky="w")

    # Results
    frm_res = ttk.Frame(root); frm_res.grid(row=3, column=0, columnspan=2, sticky="nsew"); root.rowconfigure(3, weight=1)
    cols = ("cnpj", "cnaes")
    tree = ttk.Treeview(frm_res, columns=cols, show="headings", height=12)
    tree.heading("cnpj", text="CNPJ"); tree.heading("cnaes", text="CNAEs")
    tree.column("cnpj", width=160, anchor="w"); tree.column("cnaes", width=520, anchor="w")
    yscroll = ttk.Scrollbar(frm_res, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=yscroll.set)
    tree.grid(row=0, column=0, sticky="nsew"); yscroll.grid(row=0, column=1, sticky="ns")
    frm_res.rowconfigure(0, weight=1); frm_res.columnconfigure(0, weight=1)

    # Status
    status = tk.StringVar(value="Ready")
    ttk.Label(root, textvariable=status, anchor="w").grid(row=4, column=0, columnspan=2, sticky="we", padx=6, pady=(0,6))

    # Helpers
    def _browse_file(var: tk.StringVar, filetypes):
        from tkinter import filedialog
        path = filedialog.askopenfilename(filetypes=filetypes)
        if path: var.set(path)

    def _gather_cnpjs_from_text() -> List[str]:
        raw = txt.get("1.0", "end")
        tokens = re.split(r"[\s,;]+", raw.strip())
        return _unique_cnpjs([t for t in tokens if t])

    def _gather_cnpjs_from_file() -> List[str]:
        p = Path(file_path_var.get().strip())
        if not p: raise FileNotFoundError("No file selected.")
        if not p.exists(): raise FileNotFoundError(str(p))
        ext = p.suffix.lower()
        if ext in {".csv", ".txt"}:
            sep = None if csv_sep_var.get() == "auto" else csv_sep_var.get()
            col = cnpj_col_var.get().strip() or None
            return read_cnpjs_from_csv(p, sep=sep, cnpj_col=col)
        elif ext == ".xlsx":
            col = cnpj_col_var.get().strip() or None
            sheet = xlsx_sheet_var.get().strip() or None
            return read_cnpjs_from_xlsx(p, sheet=sheet, cnpj_col=col)
        else:
            raise ValueError(f"Unsupported file type: {ext}")

    def _load_open_path() -> Path:
        p = Path(open_path_var.get().strip())
        if not p.exists() or p.stat().st_size == 0:
            raise FileNotFoundError(f"Invalid open_intervals path: {p}")
        return p

    def _set_status(msg: str):
        status.set(msg); root.update_idletasks()

    def _fill_results(df: pl.DataFrame):
        for i in tree.get_children(): tree.delete(i)
        for r in df.iter_rows(named=True):
            tree.insert("", "end", values=(r["cnpj"], ",".join(r["cnaes"])))

    def _do_lookup():
        from tkinter import messagebox
        try:
            open_path = _load_open_path()
            as_of = _validate_as_of(as_of_var.get())
            cnpjs = _gather_cnpjs_from_text() if nb.index(nb.select()) == 0 else _gather_cnpjs_from_file()
            if not cnpjs:
                messagebox.showinfo("Lookup", "No valid CNPJ found."); return
            _set_status(f"Looking up {len(cnpjs)} CNPJ(s){' as of ' + as_of if as_of else ''}...")
            if as_of:
                df = lookup_cnae_as_of_many(cnpjs, as_of, open_path=open_path, history_glob=DEFAULT_HISTORY_GLOB)
            else:
                df = lookup_cnae_open_many(cnpjs, open_path=open_path)
            _fill_results(df)
            _set_status(f"Done. {len(df)} row(s).")
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Error", str(e)); _set_status("Error")

    def _export_csv():
        from tkinter import filedialog, messagebox
        rows = [tree.item(i, "values") for i in tree.get_children()]
        if not rows:
            messagebox.showinfo("Export", "No results to export."); return
        df = pl.DataFrame({"cnpj": [r[0] for r in rows], "cnaes": [r[1].split(",") if r[1] else [] for r in rows]})
        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=(("CSV files", "*.csv"),))
        if not path: return
        export_results_csv(df, Path(path))

    root.minsize(820, 540)
    root.mainloop()


# =========== CLI ===========

def _cli():
    ap = argparse.ArgumentParser(description="CNAE lookup (current or as-of month) using open/history Parquets")
    ap.add_argument("--open-path", default=str(DEFAULT_OPEN_PATH), help="Path to open_intervals.parquet")
    ap.add_argument("--history-glob", default=DEFAULT_HISTORY_GLOB, help="Glob for history parts (closed_until_*.parquet)")
    ap.add_argument("--as-of", help="YYYY-MM for as-of lookup. Omit for current.")
    ap.add_argument("--gui", action="store_true", help="Start Tkinter GUI")
    ap.add_argument("text", nargs="*", help="CNPJ texts (CLI mode). Multiple allowed.")
    args = ap.parse_args()

    if args.gui or not args.text:
        run_gui(Path(args.open_path), args.history_glob); return

    cnpjs = _unique_cnpjs(args.text)
    if not cnpjs:
        print("No valid CNPJ found in arguments."); sys.exit(1)

    as_of = _validate_as_of(args.as_of)
    if as_of:
        df = lookup_cnae_as_of_many(cnpjs, as_of, open_path=Path(args.open_path), history_glob=args.history_glob)
    else:
        df = lookup_cnae_open_many(cnpjs, Path(args.open_path))

    for row in df.iter_rows(named=True):
        print({"cnpj": row["cnpj"], "cnaes": row["cnaes"]})


if __name__ == "__main__":
    _cli()
